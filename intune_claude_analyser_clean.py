"""
<<<<<<< HEAD
Intune AI Compliance Analyser - Excel Output with BitLocker Keys
=======
Intune AI Compliance Analyser with Excel Output
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
Uses Federated Identity OIDC No Client Secret Required
Author: Amit Bahuguna
GitHub: github.com/Amitcicd
"""

import requests
import anthropic
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from azure.identity import DefaultAzureCredential
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side



# CONFIGURATION
TENANT_ID  = os.environ.get("TENANT_ID", "your-tenant-id")
CLIENT_ID  = os.environ.get("CLIENT_ID", "your-app-client-id")
CLAUDE_API = os.environ.get("CLAUDE_API_KEY", "your-claude-api-key")

<<<<<<< HEAD
# STYLES
HEADER_FILL  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREY_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ORANGE_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
=======
# COLORS FOR EXCEL
GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE   = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

BOLD_WHITE = Font(bold=True, color="FFFFFF", size=11)
BOLD_BLACK = Font(bold=True, color="000000", size=11)
NORMAL     = Font(color="000000", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
)


def get_token():
    credential = DefaultAzureCredential(
        managed_identity_client_id=CLIENT_ID
    )
    token = credential.get_token("https://graph.microsoft.com/.default")
    return token.token


def get_headers():
    return {
        "Authorization": "Bearer " + get_token(),
        "Content-Type": "application/json"
    }


<<<<<<< HEAD
def format_date(date_str):
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str[:19], "%Y-%m-%dT%H:%M:%S")
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return date_str


=======
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
def get_all_devices():
    headers = get_headers()
    url = (
        "https://graph.microsoft.com/v1.0/deviceManagement/managedDevices"
<<<<<<< HEAD
        "?$select=id,managementAgent,ownerType,complianceState,"
        "operatingSystem,osVersion,userPrincipalName,lastSyncDateTime,"
        "isEncrypted,enrolledDateTime,manufacturer,model,joinType,"
        "azureADDeviceId,serialNumber,deviceEnrollmentType"
=======
        "?$select=managementAgent,ownerType,complianceState,"
        "operatingSystem,osVersion,userPrincipalName,lastSyncDateTime,"
        "isEncrypted,enrolledDateTime"
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    )
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json().get("value", [])


<<<<<<< HEAD
def get_bitlocker_keys():
    headers = get_headers()
    url = (
        "https://graph.microsoft.com/v1.0/informationProtection"
        "/bitlocker/recoveryKeys"
        "?$select=id,createdDateTime,deviceId,volumeType"
    )
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json().get("value", [])
    except Exception as e:
        print("BitLocker keys fetch failed: " + str(e))
        return []


def get_bitlocker_key_value(key_id):
    headers = get_headers()
    url = (
        "https://graph.microsoft.com/v1.0/informationProtection"
        "/bitlocker/recoveryKeys/" + key_id + "?$select=key"
    )
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json().get("key", "Unable to retrieve")
    except Exception:
        return "Unable to retrieve"


def get_noncompliant(devices):
    return [d for d in devices if d.get("complianceState") == "noncompliant"]


def get_stale(devices, days=30):
    cutoff = datetime.utcnow() - timedelta(days=days)
    result = []
=======
def get_noncompliant_devices(devices):
    return [d for d in devices if d.get("complianceState") == "noncompliant"]


def get_stale_devices(devices, days=30):
    cutoff = datetime.utcnow() - timedelta(days=days)
    stale = []
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    for d in devices:
        last_sync = d.get("lastSyncDateTime", "")
        if last_sync:
            try:
<<<<<<< HEAD
                sync_dt = datetime.strptime(last_sync[:19], "%Y-%m-%dT%H:%M:%S")
                if sync_dt < cutoff:
                    result.append(d)
            except Exception:
                pass
    return result


def get_unencrypted(devices):
=======
                sync_date = datetime.strptime(last_sync[:19], "%Y-%m-%dT%H:%M:%S")
                if sync_date < cutoff:
                    stale.append(d)
            except Exception:
                pass
    return stale


def get_unencrypted_devices(devices):
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    return [
        d for d in devices
        if d.get("operatingSystem") == "Windows"
        and not d.get("isEncrypted", True)
    ]
<<<<<<< HEAD
=======


def format_date(date_str):
    if not date_str:
        return "Unknown"
    try:
        dt = datetime.strptime(date_str[:19], "%Y-%m-%dT%H:%M:%S")
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return date_str
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07


def build_summary_text(devices):
    lines = []
    for d in devices:
        line = (
<<<<<<< HEAD
            "Managed By: " + str(d.get("managementAgent", "")) +
            " | Ownership: " + str(d.get("ownerType", "")) +
            " | Compliance: " + str(d.get("complianceState", "")) +
            " | OS: " + str(d.get("operatingSystem", "")) +
            " | Version: " + str(d.get("osVersion", "")) +
            " | User: " + str(d.get("userPrincipalName", "")) +
            " | Last Sync: " + format_date(d.get("lastSyncDateTime", ""))
=======
            "Managed By: " + str(d.get("managementAgent", "Unknown")) +
            " | Ownership: " + str(d.get("ownerType", "Unknown")) +
            " | Compliance: " + str(d.get("complianceState", "Unknown")) +
            " | OS: " + str(d.get("operatingSystem", "Unknown")) +
            " | Version: " + str(d.get("osVersion", "Unknown")) +
            " | User: " + str(d.get("userPrincipalName", "Unknown")) +
            " | Last Check-in: " + format_date(d.get("lastSyncDateTime", ""))
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
        )
        lines.append(line)
    return "\n".join(lines)


def analyse_with_claude(data_type, devices):
    if not devices:
        return "No issues found."
<<<<<<< HEAD
=======

>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    client = anthropic.Anthropic(api_key=CLAUDE_API)
    summary = build_summary_text(devices)

    prompts = {
        "noncompliant": (
            "You are a Microsoft Intune expert.\n"
            "Analyse these non-compliant devices:\n\n" + summary +
            "\n\nProvide:\n"
            "1. Common compliance failure patterns\n"
            "2. Priority order for remediation\n"
            "3. Specific remediation steps\n"
            "4. Devices needing immediate attention"
        ),
        "stale": (
            "You are a Microsoft Intune expert.\n"
            "Analyse these stale devices:\n\n" + summary +
            "\n\nProvide:\n"
            "1. Likely reasons for sync failure\n"
            "2. Steps to force sync\n"
            "3. Devices to consider for retirement\n"
            "4. Risk assessment"
        ),
        "bitlocker": (
            "You are a Microsoft Intune expert.\n"
            "Analyse these unencrypted Windows devices:\n\n" + summary +
            "\n\nProvide:\n"
            "1. Risk level for each device\n"
            "2. Steps to enable BitLocker via Intune\n"
            "3. Prerequisites that may block encryption\n"
<<<<<<< HEAD
            "4. Policy recommendations"
=======
            "4. Recommendations"
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
        )
    }

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1500,
        messages=[{"role": "user", "content": prompts.get(data_type, prompts["noncompliant"])}]
    )
    return message.content[0].text


<<<<<<< HEAD
def write_sheet_title(ws, title, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def write_headers(ws, headers, row=2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[row].height = 22


def auto_width(ws, headers, start_row=3):
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def create_excel_report(devices, bitlocker_keys, analyses):
    wb = openpyxl.Workbook()

    # Device lookup for BitLocker sheet
    device_lookup = {d.get("azureADDeviceId", ""): d for d in devices}

=======
def create_excel_report(devices, analyses):
    wb = openpyxl.Workbook()

>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    # ── SHEET 1: ALL DEVICES ──
    ws1 = wb.active
    ws1.title = "All Devices"

<<<<<<< HEAD
    headers1 = [
        "Managed By", "Ownership", "Compliance", "OS",
        "OS Version", "Primary User UPN", "Last Check-in",
        "Join Type", "Encrypted", "Manufacturer", "Model",
        "Enrollment Date", "Serial Number"
    ]

    write_sheet_title(ws1, "INTUNE DEVICE REPORT  |  " + datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC"), len(headers1))
    write_headers(ws1, headers1)

    for row_idx, d in enumerate(devices, 3):
        values = [
            d.get("managementAgent", ""),
            d.get("ownerType", ""),
            d.get("complianceState", ""),
            d.get("operatingSystem", ""),
            d.get("osVersion", ""),
            d.get("userPrincipalName", ""),
            format_date(d.get("lastSyncDateTime", "")),
            d.get("joinType", ""),
            "Yes" if d.get("isEncrypted") else "No",
            d.get("manufacturer", ""),
            d.get("model", ""),
            format_date(d.get("enrolledDateTime", "")),
            d.get("serialNumber", "")
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10, name="Arial")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN

            if col_idx == 3:
                compliance = str(value).lower()
                cell.fill = GREEN_FILL if compliance == "compliant" else RED_FILL if compliance == "noncompliant" else YELLOW_FILL
            elif col_idx == 9:
                cell.fill = GREEN_FILL if value == "Yes" else RED_FILL
            elif row_idx % 2 == 0:
                cell.fill = GREY_FILL

    ws1.freeze_panes = "A3"
    auto_width(ws1, headers1)

    # ── SHEET 2: NON-COMPLIANT + AI ANALYSIS ──
    ws2 = wb.create_sheet("Non-Compliant")
    noncompliant = get_noncompliant(devices)
    headers2 = ["Managed By", "Ownership", "OS", "OS Version", "Primary User UPN", "Last Check-in"]

    write_sheet_title(ws2, "NON-COMPLIANT DEVICES", len(headers2))
    write_headers(ws2, headers2)

    if noncompliant:
        for row_idx, d in enumerate(noncompliant, 3):
            values = [
                d.get("managementAgent", ""),
                d.get("ownerType", ""),
                d.get("operatingSystem", ""),
                d.get("osVersion", ""),
                d.get("userPrincipalName", ""),
                format_date(d.get("lastSyncDateTime", ""))
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(size=10, name="Arial")
                cell.fill = RED_FILL
                cell.border = THIN

        ai_row = len(noncompliant) + 4
        ws2.cell(row=ai_row, column=1, value="AI ANALYSIS AND RECOMMENDATIONS").font = Font(bold=True, name="Arial")
        analysis_cell = ws2.cell(row=ai_row + 1, column=1, value=analyses.get("noncompliant", ""))
        analysis_cell.alignment = Alignment(wrap_text=True)
        ws2.merge_cells(start_row=ai_row + 1, start_column=1, end_row=ai_row + 25, end_column=6)
        ws2.row_dimensions[ai_row + 1].height = 300
    else:
        ws2.cell(row=3, column=1, value="All devices are compliant!").font = Font(bold=True, color="00B050", name="Arial")

    auto_width(ws2, headers2)

    # ── SHEET 3: BITLOCKER KEYS ──
    ws3 = wb.create_sheet("BitLocker Keys")
    headers3 = ["Device ID", "Primary User UPN", "OS Version", "Volume Type", "Key Created Date", "Recovery Key"]

    write_sheet_title(ws3, "BITLOCKER RECOVERY KEYS", len(headers3))
    write_headers(ws3, headers3)

    if bitlocker_keys:
        for row_idx, key in enumerate(bitlocker_keys, 3):
            device_id = key.get("deviceId", "")
            device    = device_lookup.get(device_id, {})
            key_value = get_bitlocker_key_value(key.get("id", ""))

            values = [
                device_id,
                device.get("userPrincipalName", "Unknown"),
                device.get("osVersion", "Unknown"),
                key.get("volumeType", ""),
                format_date(key.get("createdDateTime", "")),
                key_value
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(size=10, name="Arial")
                cell.border = THIN
                cell.alignment = Alignment(horizontal="left")
                if col_idx == 6:
                    cell.fill = ORANGE_FILL
                elif row_idx % 2 == 0:
                    cell.fill = GREY_FILL
    else:
        ws3.cell(row=3, column=1, value="No BitLocker keys found or insufficient permissions").font = Font(name="Arial")

    ws3.freeze_panes = "A3"
    auto_width(ws3, headers3)

    # ── SHEET 4: STALE DEVICES ──
    ws4 = wb.create_sheet("Stale Devices")
    stale = get_stale(devices)
    headers4 = ["Managed By", "Ownership", "OS", "OS Version", "Primary User UPN", "Last Check-in", "Days Since Sync"]

    write_sheet_title(ws4, "STALE DEVICES (30+ DAYS NO SYNC)", len(headers4))
    write_headers(ws4, headers4)

    if stale:
        for row_idx, d in enumerate(stale, 3):
            last_sync = d.get("lastSyncDateTime", "")
            days_since = ""
            if last_sync:
                try:
                    sync_dt = datetime.strptime(last_sync[:19], "%Y-%m-%dT%H:%M:%S")
                    days_since = str((datetime.utcnow() - sync_dt).days) + " days"
                except Exception:
                    pass

            values = [
                d.get("managementAgent", ""),
                d.get("ownerType", ""),
                d.get("operatingSystem", ""),
                d.get("osVersion", ""),
                d.get("userPrincipalName", ""),
                format_date(last_sync),
                days_since
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(size=10, name="Arial")
                cell.fill = YELLOW_FILL
                cell.border = THIN
    else:
        ws4.cell(row=3, column=1, value="All devices synced recently!").font = Font(bold=True, color="00B050", name="Arial")

    auto_width(ws4, headers4)

    # ── SHEET 5: SUMMARY ──
    ws5 = wb.create_sheet("Summary")
    total       = len(devices)
    compliant   = len([d for d in devices if d.get("complianceState") == "compliant"])
    noncomp     = len(get_noncompliant(devices))
    encrypted   = len([d for d in devices if d.get("isEncrypted")])
    unencrypted = len(get_unencrypted(devices))
    stale_count = len(stale)
    corporate   = len([d for d in devices if d.get("ownerType", "").lower() == "company"])
    personal    = len([d for d in devices if d.get("ownerType", "").lower() == "personal"])

    write_sheet_title(ws5, "COMPLIANCE SUMMARY DASHBOARD", 3)
    write_headers(ws5, ["Metric", "Count", "Status"], row=2)

    summary_data = [
        ("Total Devices",           total,       ""),
        ("Compliant",               compliant,   "Healthy"),
        ("Non-Compliant",           noncomp,     "Action Required" if noncomp > 0 else "Healthy"),
        ("Encrypted (Windows)",     encrypted,   "Healthy"),
        ("Unencrypted (Windows)",   unencrypted, "Action Required" if unencrypted > 0 else "Healthy"),
        ("Stale Devices (30d)",     stale_count, "Review Required" if stale_count > 0 else "Healthy"),
        ("Corporate Owned",         corporate,   ""),
        ("Personal/BYOD",           personal,    ""),
        ("BitLocker Keys Found",    len(bitlocker_keys), ""),
    ]

    fills5 = [GREY_FILL, GREEN_FILL, RED_FILL if noncomp > 0 else GREEN_FILL,
              GREEN_FILL, RED_FILL if unencrypted > 0 else GREEN_FILL,
              YELLOW_FILL if stale_count > 0 else GREEN_FILL,
              GREY_FILL, GREY_FILL, ORANGE_FILL]

    for row_idx, ((metric, count, status), fill) in enumerate(zip(summary_data, fills5), 3):
        ws5.cell(row=row_idx, column=1, value=metric).border = THIN
        ws5.cell(row=row_idx, column=1).font = Font(name="Arial", size=10)
        count_cell = ws5.cell(row=row_idx, column=2, value=count)
        count_cell.border = THIN
        count_cell.font = Font(name="Arial", size=10, bold=True)
        count_cell.fill = fill
        count_cell.alignment = Alignment(horizontal="center")
        ws5.cell(row=row_idx, column=3, value=status).border = THIN
        ws5.cell(row=row_idx, column=3).font = Font(name="Arial", size=10)

    for col_letter, width in [("A", 28), ("B", 12), ("C", 20)]:
        ws5.column_dimensions[col_letter].width = width

    # Save
    filename = "Intune_Report_" + datetime.utcnow().strftime("%Y%m%d_%H%M") + ".xlsx"
=======
    # Title row
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "INTUNE DEVICE COMPLIANCE REPORT - " + datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    ws1["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws1["A1"].fill = HEADER
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Header row
    headers = [
        "Managed By", "Ownership", "Compliance",
        "OS", "OS Version", "Primary User UPN",
        "Last Check-in", "Encrypted"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = BOLD_WHITE
        cell.fill = HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Device rows
    for row, d in enumerate(devices, 3):
        compliance = d.get("complianceState", "Unknown")
        encrypted  = d.get("isEncrypted", True)

        values = [
            d.get("managementAgent", "Unknown"),
            d.get("ownerType", "Unknown"),
            compliance,
            d.get("operatingSystem", "Unknown"),
            d.get("osVersion", "Unknown"),
            d.get("userPrincipalName", "Unknown"),
            format_date(d.get("lastSyncDateTime", "")),
            "Yes" if encrypted else "No"
        ]

        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.font = NORMAL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left")

            # Color compliance column
            if col == 3:
                if compliance == "compliant":
                    cell.fill = GREEN
                elif compliance == "noncompliant":
                    cell.fill = RED
                else:
                    cell.fill = YELLOW

            # Color encrypted column
            if col == 8:
                if not encrypted and d.get("operatingSystem") == "Windows":
                    cell.fill = RED
                else:
                    cell.fill = GREEN

    # Auto column width
    for col in ws1.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── SHEET 2: NON-COMPLIANT ──
    ws2 = wb.create_sheet("Non-Compliant")
    ws2["A1"] = "NON-COMPLIANT DEVICES ANALYSIS"
    ws2["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws2["A1"].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws2.merge_cells("A1:H1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    noncompliant = [d for d in devices if d.get("complianceState") == "noncompliant"]
    if noncompliant:
        headers2 = ["Managed By", "Ownership", "OS", "OS Version", "Primary User UPN", "Last Check-in"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=2, column=col, value=header)
            cell.font = BOLD_WHITE
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        for row, d in enumerate(noncompliant, 3):
            values = [
                d.get("managementAgent", "Unknown"),
                d.get("ownerType", "Unknown"),
                d.get("operatingSystem", "Unknown"),
                d.get("osVersion", "Unknown"),
                d.get("userPrincipalName", "Unknown"),
                format_date(d.get("lastSyncDateTime", ""))
            ]
            for col, value in enumerate(values, 1):
                cell = ws2.cell(row=row, column=col, value=value)
                cell.font = NORMAL
                cell.fill = RED
                cell.border = THIN_BORDER

        # AI Analysis
        analysis_row = len(noncompliant) + 4
        ws2.cell(row=analysis_row, column=1, value="AI ANALYSIS AND RECOMMENDATIONS")
        ws2.cell(row=analysis_row, column=1).font = BOLD_BLACK
        ws2.cell(row=analysis_row + 1, column=1, value=analyses.get("noncompliant", "No analysis available"))
        ws2.cell(row=analysis_row + 1, column=1).alignment = Alignment(wrap_text=True)
        ws2.merge_cells(
            start_row=analysis_row + 1, start_column=1,
            end_row=analysis_row + 20, end_column=6
        )
    else:
        ws2["A2"] = "All devices are compliant!"
        ws2["A2"].font = Font(bold=True, color="00B050", size=12)

    # Auto width sheet 2
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── SHEET 3: SUMMARY ──
    ws3 = wb.create_sheet("Summary")
    ws3["A1"] = "COMPLIANCE SUMMARY"
    ws3["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws3["A1"].fill = HEADER
    ws3.merge_cells("A1:C1")
    ws3["A1"].alignment = Alignment(horizontal="center")

    total       = len(devices)
    compliant   = len([d for d in devices if d.get("complianceState") == "compliant"])
    noncompliant_count = len([d for d in devices if d.get("complianceState") == "noncompliant"])
    unencrypted = len([d for d in devices if d.get("operatingSystem") == "Windows" and not d.get("isEncrypted", True)])
    stale_count = len(get_stale_devices(devices))

    summary_data = [
        ("Total Devices", total, ""),
        ("Compliant", compliant, str(round(compliant/total*100 if total else 0, 1)) + "%"),
        ("Non-Compliant", noncompliant_count, str(round(noncompliant_count/total*100 if total else 0, 1)) + "%"),
        ("Unencrypted Windows", unencrypted, "Needs attention"),
        ("Stale Devices (30d)", stale_count, "Review recommended"),
    ]

    ws3.cell(row=2, column=1, value="Metric").font = BOLD_BLACK
    ws3.cell(row=2, column=2, value="Count").font = BOLD_BLACK
    ws3.cell(row=2, column=3, value="Details").font = BOLD_BLACK

    for row, (metric, count, detail) in enumerate(summary_data, 3):
        ws3.cell(row=row, column=1, value=metric).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=count).border = THIN_BORDER
        ws3.cell(row=row, column=3, value=detail).border = THIN_BORDER

        if metric == "Compliant":
            ws3.cell(row=row, column=2).fill = GREEN
        elif metric in ("Non-Compliant", "Unencrypted Windows"):
            ws3.cell(row=row, column=2).fill = RED
        elif metric == "Stale Devices (30d)":
            ws3.cell(row=row, column=2).fill = YELLOW

    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 30

    # Save
    filename = "intune_report_" + datetime.utcnow().strftime("%Y%m%d_%H%M") + ".xlsx"
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07
    wb.save(filename)
    return filename


def main():
<<<<<<< HEAD
    print("Authenticating...")
    print("Fetching devices from Intune...")
    devices = get_all_devices()
    print("Found " + str(len(devices)) + " devices")

    print("Fetching BitLocker recovery keys...")
    bitlocker_keys = get_bitlocker_keys()
    print("Found " + str(len(bitlocker_keys)) + " BitLocker keys")

    print("Running AI analysis...")
    analyses = {}
    noncompliant = get_noncompliant(devices)
    stale        = get_stale(devices)
    unencrypted  = get_unencrypted(devices)

    analyses["noncompliant"] = analyse_with_claude("noncompliant", noncompliant) if noncompliant else "All devices compliant!"
    analyses["stale"]        = analyse_with_claude("stale", stale) if stale else "All devices synced recently!"
    analyses["bitlocker"]    = analyse_with_claude("bitlocker", unencrypted) if unencrypted else "All Windows devices encrypted!"

    print("Generating Excel report...")
    filename = create_excel_report(devices, bitlocker_keys, analyses)
    print("Report saved: " + filename)
    print("\nSUMMARY")
    print("Total devices:   " + str(len(devices)))
    print("Non-compliant:   " + str(len(noncompliant)))
    print("Stale devices:   " + str(len(stale)))
    print("Unencrypted:     " + str(len(unencrypted)))
    print("BitLocker keys:  " + str(len(bitlocker_keys)))
=======
    print("Authenticating with federated identity...")

    print("Fetching all devices from Intune...")
    devices = get_all_devices()
    print("Found " + str(len(devices)) + " total devices")

    noncompliant = get_noncompliant_devices(devices)
    stale        = get_stale_devices(devices)
    unencrypted  = get_unencrypted_devices(devices)

    print("Running AI analysis...")
    analyses = {}

    if noncompliant:
        print("Analysing " + str(len(noncompliant)) + " non-compliant devices...")
        analyses["noncompliant"] = analyse_with_claude("noncompliant", noncompliant)
    else:
        analyses["noncompliant"] = "All devices are compliant!"

    if stale:
        print("Analysing " + str(len(stale)) + " stale devices...")
        analyses["stale"] = analyse_with_claude("stale", stale)
    else:
        analyses["stale"] = "All devices synced recently!"

    if unencrypted:
        print("Analysing " + str(len(unencrypted)) + " unencrypted devices...")
        analyses["bitlocker"] = analyse_with_claude("bitlocker", unencrypted)
    else:
        analyses["bitlocker"] = "All Windows devices are encrypted!"

    print("Generating Excel report...")
    filename = create_excel_report(devices, analyses)
    print("Report saved to: " + filename)
>>>>>>> cc5d34da88e308e433fba202599bbd9c6c85ec07

    print("\nSUMMARY:")
    print("Total devices: " + str(len(devices)))
    print("Non-compliant: " + str(len(noncompliant)))
    print("Stale devices: " + str(len(stale)))
    print("Unencrypted:   " + str(len(unencrypted)))


if __name__ == "__main__":
    main()
